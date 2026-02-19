import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import config, logger
from utils.normalization import normalize_vessel_name

BASE_DIR = Path(__file__).resolve().parent.parent
FUZZY_THRESHOLD = config["processor"].get("fuzzy_match_threshold", 85)

# Regex for dates like "13.02.2026" or "13.02.2026 05:30"
DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


def parse_eta_date(eta_str: str) -> str | None:
    """Extract date part (DD.MM.YYYY) from an ETA string."""
    m = DATE_RE.search(str(eta_str))
    return m.group(0) if m else None


def cross_match(
    eurogate: list[dict],
    hhla: list[dict],
    threshold: int = FUZZY_THRESHOLD,
) -> list[dict]:
    """Match Eurogate vessels against HHLA to enrich data.

    Only matches ACROSS sources (Eurogate <-> HHLA), never within.
    A match requires: fuzzy name score >= threshold AND same ETA date.
    """
    result = []
    hhla_used = set()

    for eg in eurogate:
        eg_name = normalize_vessel_name(eg.get("vessel_name", ""))
        eg_date = parse_eta_date(eg.get("eta", ""))

        best_match = None
        best_score = 0

        for j, hh in enumerate(hhla):
            if j in hhla_used:
                continue

            hh_name = normalize_vessel_name(hh.get("vessel_name", ""))
            score = fuzz.ratio(eg_name, hh_name)

            if score < threshold:
                continue

            # Must also have matching ETA date
            hh_date = parse_eta_date(hh.get("eta", ""))
            if eg_date and hh_date and eg_date != hh_date:
                continue

            if score > best_score:
                best_score = score
                best_match = j

        if best_match is not None:
            hhla_used.add(best_match)
            merged = _merge_pair(eg, hhla[best_match])
            merged["match_score"] = best_score
            result.append(merged)
        else:
            eg_copy = dict(eg)
            eg_copy["match_score"] = 0
            result.append(eg_copy)

    # Add unmatched HHLA vessels
    for j, hh in enumerate(hhla):
        if j not in hhla_used:
            hh_copy = dict(hh)
            hh_copy["match_score"] = 0
            result.append(hh_copy)

    return result


def _merge_pair(eg: dict, hh: dict) -> dict:
    """Merge one Eurogate and one HHLA record."""
    merged = dict(eg)
    terminals = {eg.get("terminal", ""), hh.get("terminal", "")}
    merged["terminal"] = " + ".join(sorted(t for t in terminals if t))

    # Fill empty fields from HHLA
    for key, val in hh.items():
        if val and not merged.get(key):
            merged[key] = val

    return merged


def build_dataframe(vessels: list[dict]) -> pd.DataFrame:
    """Convert vessel list to a clean DataFrame."""
    df = pd.DataFrame(vessels)

    # Replace NaN with empty strings
    df = df.fillna("")

    # Parse ETA for sorting: extract DD.MM.YYYY HH:MM and convert
    def sort_key(eta_str):
        m = DATE_RE.search(str(eta_str))
        if not m:
            return "9999"
        d, mo, y = m.groups()
        # Also grab time if present
        time_part = "00:00"
        rest = str(eta_str)[m.end():].strip()
        if rest and ":" in rest:
            time_part = rest[:5]
        return f"{y}-{mo}-{d} {time_part}"

    df["_sort"] = df["eta"].apply(sort_key)
    df = df.sort_values("_sort").reset_index(drop=True)
    df = df.drop(columns=["_sort"])

    # Add metadata
    df["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    return df


def export_excel(df: pd.DataFrame, filename: str | None = None) -> Path:
    """Export DataFrame to formatted Excel file."""
    if filename is None:
        filename = config["processor"]["excel"].get(
            "output_file", "data/vessel_schedule.xlsx"
        )

    out_path = BASE_DIR / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Vessels", index=False)
        ws = writer.sheets["Vessels"]

        # Style header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit column widths
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            col_name = df.columns[col_idx - 1]
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

    logger.info(f"[processor] Excel exported: {out_path.name} ({len(df)} rows)")
    return out_path


def process(eurogate_vessels: list[dict], hhla_vessels: list[dict], output_path: str | None = None) -> Path:
    """Main processing pipeline: cross-match, merge, export."""
    logger.info(
        f"[processor] Input: {len(eurogate_vessels)} Eurogate + {len(hhla_vessels)} HHLA"
    )

    # Cross-match between sources (not within)
    all_vessels = cross_match(eurogate_vessels, hhla_vessels)

    matched = sum(1 for v in all_vessels if v.get("match_score", 0) > 0)
    logger.info(
        f"[processor] Result: {len(all_vessels)} total, "
        f"{matched} cross-matched between terminals"
    )

    df = build_dataframe(all_vessels)
    out_path = export_excel(df, filename=output_path)

    return out_path
